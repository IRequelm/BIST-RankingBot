from __future__ import annotations

import json
from pathlib import Path
from shutil import copyfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.current_portfolio import _current_regime, _latest_feature_snapshot, _rank_snapshot


START_VALUE = 100_000.0
TRACKING_STATE_FILENAME = "tracking_state.json"


def _price_at_or_before(prices: pd.DataFrame, date: pd.Timestamp) -> tuple[pd.Timestamp, float]:
    close = prices["Close"].dropna().sort_index()
    available = close[close.index <= date]
    if available.empty:
        available = close
    return pd.Timestamp(available.index[-1]), float(available.iloc[-1])


def _latest_price(prices: pd.DataFrame) -> tuple[pd.Timestamp, float]:
    close = prices["Close"].dropna().sort_index()
    return pd.Timestamp(close.index[-1]), float(close.iloc[-1])


def _state_path(reports_path: Path) -> Path:
    return reports_path / TRACKING_STATE_FILENAME


def _load_tracking_state(path: Path) -> dict[str, object] | None:
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as handle:
        state = json.load(handle)
    if not state.get("positions"):
        raise ValueError(f"Tracking state has no positions: {path}")
    return state


def _save_tracking_state(path: Path, state: dict[str, object]) -> None:
    path.write_text(json.dumps(state, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _candidate_selection(
    stock_prices: dict[str, pd.DataFrame],
    benchmark_prices: pd.DataFrame,
    factor_models: dict[str, dict[str, float]],
) -> tuple[pd.DataFrame, dict[str, object], pd.Timestamp]:
    regime = _current_regime(benchmark_prices)
    snapshot, snapshot_date = _latest_feature_snapshot(stock_prices)
    if snapshot.empty:
        raise ValueError("No valid stock features available for one-month follow-up report.")

    active_model = "low_volatility" if regime["bist100_below_ma200"] else "volume_heavy"
    ranked = _rank_snapshot(snapshot, factor_models[active_model], active_model).head(3).copy()
    ranked["weight"] = 1.0 / len(ranked)
    ranked["active_model"] = active_model
    return ranked, regime, pd.Timestamp(snapshot_date)


def _create_tracking_state(
    selected: pd.DataFrame,
    stock_prices: dict[str, pd.DataFrame],
    benchmark_prices: pd.DataFrame,
    start_date: pd.Timestamp,
    regime: dict[str, object],
) -> dict[str, object]:
    positions = []
    for _, row in selected.iterrows():
        symbol = row["symbol"]
        price_date, start_price = _price_at_or_before(stock_prices[symbol], start_date)
        positions.append(
            {
                "symbol": symbol,
                "weight": float(row["weight"]),
                "start_date": price_date.strftime("%Y-%m-%d"),
                "start_price": start_price,
            }
        )

    benchmark_start_date, benchmark_start_price = _price_at_or_before(benchmark_prices, start_date)
    active_model = str(selected["active_model"].iloc[0]) if "active_model" in selected else "unknown"
    return {
        "version": 1,
        "strategy": "Top3 Ranking Only",
        "tracking_mode": "fixed_portfolio",
        "created_at": pd.Timestamp.today().normalize().strftime("%Y-%m-%d"),
        "start_date": pd.Timestamp(start_date).strftime("%Y-%m-%d"),
        "start_value": START_VALUE,
        "active_model": active_model,
        "regime_status_at_start": regime["regime_status"],
        "benchmark": {
            "symbol": "XU100.IS",
            "start_date": benchmark_start_date.strftime("%Y-%m-%d"),
            "start_price": benchmark_start_price,
        },
        "positions": positions,
    }


def _state_to_selected(state: dict[str, object]) -> pd.DataFrame:
    positions = pd.DataFrame(state["positions"])
    if positions.empty:
        raise ValueError("Tracking state positions are empty.")
    positions["active_model"] = state.get("active_model", "unknown")
    return positions


def _build_positions(
    state: dict[str, object],
    stock_prices: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    rows = []
    for position in state["positions"]:
        symbol = position["symbol"]
        if symbol not in stock_prices:
            raise ValueError(f"Tracked symbol is missing from loaded prices: {symbol}")

        start_price = float(position["start_price"])
        current_date, current_price = _latest_price(stock_prices[symbol])
        weight = float(position["weight"])
        start_amount = START_VALUE * weight
        current_amount = start_amount * (current_price / start_price) if start_price else start_amount
        rows.append(
            {
                "Hisse": symbol,
                "Ağırlık %": weight,
                "Başlangıç Fiyatı": start_price,
                "Güncel Fiyat": current_price,
                "Getiri %": (current_price / start_price) - 1 if start_price else 0.0,
                "Başlangıç Tutarı": start_amount,
                "Güncel Tutar": current_amount,
                "Kar/Zarar TL": current_amount - start_amount,
                "_current_date": current_date,
            }
        )
    return pd.DataFrame(rows)


def _build_daily_tracking(
    selected: pd.DataFrame,
    stock_prices: dict[str, pd.DataFrame],
    benchmark_prices: pd.DataFrame,
    state: dict[str, object],
) -> pd.DataFrame:
    start_date = pd.Timestamp(state["start_date"])
    curves = []
    for _, row in selected.iterrows():
        symbol = row["symbol"]
        if symbol not in stock_prices:
            raise ValueError(f"Tracked symbol is missing from loaded prices: {symbol}")

        close = stock_prices[symbol]["Close"].dropna().sort_index()
        start_price = float(row["start_price"])
        window = close[close.index >= start_date].copy()
        if window.empty:
            window = close.tail(1).copy()
        curve = (window / start_price) * float(row["weight"])
        curve.loc[start_date] = float(row["weight"])
        curves.append(curve.sort_index())

    portfolio_curve = pd.concat(curves, axis=1).sort_index().ffill().bfill().sum(axis=1)
    benchmark_close = benchmark_prices["Close"].dropna().sort_index()
    benchmark_state = state.get("benchmark", {})
    benchmark_start = float(benchmark_state.get("start_price", 0.0))
    if not benchmark_start:
        _, benchmark_start = _price_at_or_before(benchmark_prices, start_date)
    benchmark_window = benchmark_close[benchmark_close.index >= start_date].copy()
    if benchmark_window.empty:
        benchmark_window = benchmark_close.tail(1).copy()
    benchmark_curve = benchmark_window / benchmark_start
    benchmark_curve.loc[start_date] = 1.0
    benchmark_curve = benchmark_curve.sort_index()

    tracking = pd.concat(
        [
            (portfolio_curve * START_VALUE).rename("Portföy Değeri"),
            portfolio_curve.rename("_portfolio_curve"),
            benchmark_window.rename("BIST100 Değeri"),
            benchmark_curve.rename("_benchmark_curve"),
        ],
        axis=1,
    ).sort_index().ffill().dropna()
    tracking["Günlük Getiri %"] = tracking["_portfolio_curve"].pct_change().fillna(0.0)
    tracking["Toplam Getiri %"] = tracking["_portfolio_curve"] - 1
    tracking["BIST100 Getiri %"] = tracking["_benchmark_curve"] - 1
    tracking["Fark %"] = tracking["Toplam Getiri %"] - tracking["BIST100 Getiri %"]
    tracking = tracking.reset_index().rename(columns={"Date": "Tarih", "index": "Tarih"})
    return tracking[
        [
            "Tarih",
            "Portföy Değeri",
            "Günlük Getiri %",
            "Toplam Getiri %",
            "BIST100 Değeri",
            "BIST100 Getiri %",
            "Fark %",
        ]
    ]


def _style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 38)

    percent_headers = {"Ağırlık %", "Getiri %", "Günlük Getiri %", "Toplam Getiri %", "BIST100 Getiri %", "Fark %"}
    money_headers = {
        "Başlangıç Fiyatı",
        "Güncel Fiyat",
        "Başlangıç Tutarı",
        "Güncel Tutar",
        "Kar/Zarar TL",
        "Portföy Değeri",
        "BIST100 Değeri",
    }
    for sheet in workbook.worksheets:
        headers = {cell.value: cell.column for cell in sheet[1]}
        for header in percent_headers & headers.keys():
            for col in sheet.iter_cols(min_col=headers[header], max_col=headers[header], min_row=2):
                for cell in col:
                    cell.number_format = "0.00%"
        for header in money_headers & headers.keys():
            for col in sheet.iter_cols(min_col=headers[header], max_col=headers[header], min_row=2):
                for cell in col:
                    cell.number_format = "#,##0.00"
    workbook.save(path)


def generate_one_month_follow_up(
    stock_prices: dict[str, pd.DataFrame],
    benchmark_prices: pd.DataFrame,
    factor_models: dict[str, dict[str, float]],
    reports_dir: str = "reports",
) -> tuple[Path, Path]:
    reports_path = Path(reports_dir)
    reports_path.mkdir(parents=True, exist_ok=True)

    current_regime = _current_regime(benchmark_prices)
    state_file = _state_path(reports_path)
    state = _load_tracking_state(state_file)
    if state is None:
        selected, start_regime, start_date = _candidate_selection(stock_prices, benchmark_prices, factor_models)
        state = _create_tracking_state(selected, stock_prices, benchmark_prices, start_date, start_regime)
        _save_tracking_state(state_file, state)

    selected = _state_to_selected(state)
    positions = _build_positions(state, stock_prices)
    daily = _build_daily_tracking(selected, stock_prices, benchmark_prices, state)

    portfolio_current_value = float(positions["Güncel Tutar"].sum())
    portfolio_return = (portfolio_current_value / START_VALUE) - 1
    start_date = pd.Timestamp(state["start_date"])
    report_date = pd.Timestamp(max(positions["_current_date"]))
    benchmark_state = state.get("benchmark", {})
    benchmark_start_date = pd.Timestamp(benchmark_state.get("start_date", start_date))
    benchmark_start = float(benchmark_state.get("start_price", 0.0))
    if not benchmark_start:
        benchmark_start_date, benchmark_start = _price_at_or_before(benchmark_prices, start_date)
    benchmark_current_date, benchmark_current = _latest_price(benchmark_prices)
    benchmark_return = (benchmark_current / benchmark_start) - 1 if benchmark_start else 0.0
    excess_return = portfolio_return - benchmark_return
    status = f"{'Karda' if portfolio_return >= 0 else 'Zararda'} / {'BIST Üstü' if excess_return >= 0 else 'BIST Altı'}"

    summary = pd.DataFrame(
        [
            {"Alan": "Rapor Tarihi", "Değer": report_date.strftime("%Y-%m-%d")},
            {"Alan": "Takip Modu", "Değer": "Sabit portföy"},
            {"Alan": "Strateji", "Değer": state.get("strategy", "Top3 Ranking Only")},
            {"Alan": "Başlangıç Modeli", "Değer": state.get("active_model", "unknown")},
            {"Alan": "Başlangıç Rejim Durumu", "Değer": state.get("regime_status_at_start", "unknown")},
            {"Alan": "Güncel Rejim Durumu", "Değer": current_regime["regime_status"]},
            {"Alan": "BIST100 Kapanış", "Değer": float(current_regime["bist100_close"])},
            {"Alan": "Portföy Başlangıç Değeri", "Değer": START_VALUE},
            {"Alan": "Seçilen 3 Hisse", "Değer": ", ".join(positions["Hisse"])},
            {"Alan": "Eşit Ağırlık", "Değer": "33.33% / 33.33% / 33.33%"},
            {"Alan": "Başlangıç Tarihi", "Değer": start_date.strftime("%Y-%m-%d")},
            {"Alan": "Güncel Fiyat Tarihi", "Değer": max(positions["_current_date"]).strftime("%Y-%m-%d")},
            {"Alan": "BIST100 Başlangıç Tarihi", "Değer": benchmark_start_date.strftime("%Y-%m-%d")},
            {"Alan": "BIST100 Güncel Tarihi", "Değer": benchmark_current_date.strftime("%Y-%m-%d")},
            {"Alan": "Portföy Toplam Getiri", "Değer": portfolio_return},
            {"Alan": "BIST100 Getiri", "Değer": benchmark_return},
            {"Alan": "BIST100'e Göre Fark", "Değer": excess_return},
            {"Alan": "Durum", "Değer": status},
        ]
    )

    notes = pd.DataFrame(
        {
            "Notlar": [
                "Bu dosya araştırma dosyası değildir.",
                "Bir aylık takip dosyasıdır.",
                "Alım/satım önerisi değildir.",
                "Amaç stratejinin BIST100'e göre performansını takip etmektir.",
                "Portföy reports/tracking_state.json içinden okunur.",
                "Sonraki çalıştırmalarda yeni hisse seçilmez.",
            ]
        }
    )

    positions_public = positions.drop(columns=["_current_date"])
    xlsx_path = reports_path / "one_month_follow_up.xlsx"
    md_path = reports_path / "one_month_follow_up.md"
    latest_xlsx_path = reports_path / "latest_one_month_follow_up.xlsx"
    latest_md_path = reports_path / "latest_one_month_follow_up.md"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="TAKIP_OZETI", index=False)
        positions_public.to_excel(writer, sheet_name="POZISYONLAR", index=False)
        daily.to_excel(writer, sheet_name="GUNLUK_TAKIP", index=False)
        notes.to_excel(writer, sheet_name="NOTLAR", index=False)
    _style_workbook(xlsx_path)

    lines = [
        "# Bir Aylık Takip Raporu",
        "",
        f"- Rapor Tarihi: {report_date.strftime('%Y-%m-%d')}",
        "- Takip Modu: Sabit portföy",
        f"- Başlangıç Tarihi: {start_date.strftime('%Y-%m-%d')}",
        f"- Strateji: {state.get('strategy', 'Top3 Ranking Only')}",
        f"- Başlangıç Modeli: {state.get('active_model', 'unknown')}",
        f"- Başlangıç Rejim Durumu: {state.get('regime_status_at_start', 'unknown')}",
        f"- Güncel Rejim Durumu: {current_regime['regime_status']}",
        f"- Başlangıç Değeri: {START_VALUE:,.2f} TL",
        f"- Seçilen Hisseler: {', '.join(positions['Hisse'])}",
        f"- Portföy Toplam Getiri: {portfolio_return:.2%}",
        f"- BIST100 Getiri: {benchmark_return:.2%}",
        f"- BIST100'e Göre Fark: {excess_return:.2%}",
        f"- Durum: {status}",
        "",
        "## Pozisyonlar",
        "",
        positions_public.to_markdown(index=False, floatfmt=".4f"),
        "",
        "## Notlar",
        "",
        "- Bu dosya araştırma dosyası değildir.",
        "- Bir aylık takip dosyasıdır.",
        "- Alım/satım önerisi değildir.",
        "- Amaç stratejinin BIST100'e göre performansını takip etmektir.",
        "- Portföy reports/tracking_state.json içinden okunur.",
        "- Sonraki çalıştırmalarda yeni hisse seçilmez.",
        "",
    ]
    md_path.write_text("\n".join(lines), encoding="utf-8")
    copyfile(xlsx_path, latest_xlsx_path)
    copyfile(md_path, latest_md_path)
    return xlsx_path, md_path
