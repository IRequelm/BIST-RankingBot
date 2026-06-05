from datetime import datetime
from pathlib import Path
from shutil import copyfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ACTION_LABELS = {
    "BUY": "AL",
    "HOLD": "TUT",
    "SELL": "SAT",
    "EXCLUDE": "PORTFOY_DISI",
}

ACTION_FILLS = {
    "AL": "D9EAD3",
    "TUT": "FFF2CC",
    "SAT": "F4CCCC",
    "PORTFOY_DISI": "E7E6E6",
    "CASH": "D9EAF7",
}

REPORT_CAPITAL = 100_000.0


def _latest_close(stock_prices: dict[str, pd.DataFrame], symbol: str) -> float | None:
    prices = stock_prices.get(symbol)
    if prices is None or prices.empty or "Close" not in prices:
        return None

    close = prices["Close"].dropna()
    if close.empty:
        return None
    return float(close.iloc[-1])


def _action_label(action: object) -> str:
    return ACTION_LABELS.get(str(action), str(action))


def _price_target(entry_price: float | None, expected_return: float | None) -> float | None:
    if entry_price is None or pd.isna(entry_price) or expected_return is None or pd.isna(expected_return):
        return None
    return entry_price * (1 + float(expected_return))


def _risk_price(entry_price: float | None, expected_low: float | None, volatility: float | None) -> float | None:
    if entry_price is None or pd.isna(entry_price):
        return None

    candidates = []
    if expected_low is not None and not pd.isna(expected_low):
        candidates.append(entry_price * (1 + float(expected_low)))
    if volatility is not None and not pd.isna(volatility):
        candidates.append(entry_price * (1 - max(float(volatility) * 2, 0.05)))

    return min(candidates) if candidates else entry_price * 0.92


def _risk_reward(entry_price: float | None, target_price: float | None, stop_price: float | None) -> float | None:
    if (
        entry_price is None
        or target_price is None
        or stop_price is None
        or pd.isna(entry_price)
        or pd.isna(target_price)
        or pd.isna(stop_price)
    ):
        return None

    upside = max(float(target_price) - float(entry_price), 0)
    downside = max(float(entry_price) - float(stop_price), 0)
    if downside == 0:
        return None
    return upside / downside


def _conviction_level(score: float | None, expected_mid: float | None, risk_reward: float | None) -> str:
    score_value = 0.0 if score is None or pd.isna(score) else float(score)
    expected_value = 0.0 if expected_mid is None or pd.isna(expected_mid) else float(expected_mid)
    rr_value = 0.0 if risk_reward is None or pd.isna(risk_reward) else float(risk_reward)

    conviction_score = score_value * 0.55 + min(max(expected_value, 0), 0.20) * 2.0 + min(rr_value, 3.0) * 0.10
    if conviction_score >= 0.80:
        return "VERY HIGH"
    if conviction_score >= 0.62:
        return "HIGH"
    if conviction_score >= 0.45:
        return "MEDIUM"
    return "LOW"


def _attractiveness_score(row: pd.Series, risk_reward: float | None) -> float:
    expected_mid = row.get("expected_return_mid")
    expected_value = 0.0 if expected_mid is None or pd.isna(expected_mid) else float(expected_mid)
    rr_value = 0.0 if risk_reward is None or pd.isna(risk_reward) else float(risk_reward)
    volatility = row.get("volatility")
    volatility_penalty = 0.0 if volatility is None or pd.isna(volatility) else float(volatility)
    return float(row.get("score", 0)) + expected_value + min(rr_value, 3.0) * 0.08 - volatility_penalty


def _trend_label(above_ma: object) -> str:
    if pd.isna(above_ma):
        return "Belirsiz"
    value = float(above_ma)
    if value >= 0.75:
        return "Güçlü trend"
    if value >= 0.50:
        return "Kısmen trend üstü"
    if value > 0:
        return "Zayıf trend"
    return "Trend altı"


def _main_reason(row: pd.Series) -> str:
    parts = []
    if row.get("momentum_6m", 0) > 0:
        parts.append("6A momentum pozitif")
    if row.get("momentum_3m", 0) > 0:
        parts.append("3A momentum destekli")
    if row.get("above_ma", 0) >= 0.5:
        parts.append("trend desteği var")
    if row.get("volume_increase", 0) > 0:
        parts.append("hacim artışı var")
    if row.get("volatility", 1) < 0.025:
        parts.append("volatilite görece düşük")

    if not parts:
        return "Model skoru portföy için yeterli"
    return "; ".join(parts[:3])


def _risk_note(row: pd.Series) -> str:
    notes = []
    if row.get("momentum_1m", 0) < 0:
        notes.append("kisa vadeli momentum negatif")
    if row.get("above_ma", 0) < 0.5:
        notes.append("trend desteği zayıf")
    if row.get("volatility", 0) >= 0.03:
        notes.append("volatilite yüksek")
    if row.get("expected_return_low", 0) < -0.05:
        notes.append("alt bant riski belirgin")

    if not notes:
        return "Risk sinyali sınırlı"
    return "; ".join(notes[:3])


def _load_csv(path: Path) -> pd.DataFrame:
    if path.exists():
        return pd.read_csv(path)
    return pd.DataFrame()


def _as_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "evet"}


def _report_date_from_current(current: pd.DataFrame) -> str:
    for column in ["snapshot_date", "benchmark_date"]:
        if column in current and pd.notna(current[column].iloc[0]):
            return pd.Timestamp(current[column].iloc[0]).strftime("%Y-%m-%d")
    return datetime.now().strftime("%Y-%m-%d")


def _report_is_stale(report_date: str, max_age_days: int = 35) -> bool:
    age_days = (pd.Timestamp(datetime.now().date()) - pd.Timestamp(report_date)).days
    return age_days > max_age_days


def _cash_weight(executive_summary: pd.DataFrame) -> float:
    if executive_summary.empty:
        return 0.0
    cash = executive_summary.loc[executive_summary["Stock"] == "CASH", "Weight %"]
    return float(cash.iloc[0]) if not cash.empty and pd.notna(cash.iloc[0]) else 0.0


def _quality_for_row(row: pd.Series, entry_price: float | None = None) -> str:
    expected_mid = row.get("expected_return_mid")
    target_price = _price_target(entry_price, expected_mid) if entry_price is not None else None
    stop_price = _risk_price(entry_price, row.get("expected_return_low"), row.get("volatility")) if entry_price is not None else None
    return _conviction_level(row.get("score"), expected_mid, _risk_reward(entry_price, target_price, stop_price))


def _build_executive_summary(current: pd.DataFrame, stock_prices: dict[str, pd.DataFrame]) -> pd.DataFrame:
    recommended = current[current["action"] == "BUY"].copy()
    rows = []
    active_portfolio_size = int(current["active_portfolio_size"].iloc[0]) if "active_portfolio_size" in current else 10

    for _, row in recommended.iterrows():
        symbol = row["symbol"]
        entry_price = _latest_close(stock_prices, symbol)
        expected_mid = row.get("expected_return_mid")
        target_price = _price_target(entry_price, expected_mid)
        stop_price = _risk_price(entry_price, row.get("expected_return_low"), row.get("volatility"))
        rr = _risk_reward(entry_price, target_price, stop_price)
        attractiveness = _attractiveness_score(row, rr)
        rows.append(
            {
                "Stock": symbol,
                "Action": _action_label(row["action"]),
                "Expected Return %": expected_mid,
                "Target Price": target_price,
                "Stop Price": stop_price,
                "Risk/Reward": rr,
                "Conviction": _conviction_level(row.get("score"), expected_mid, rr),
                "_attractiveness": attractiveness,
            }
        )

    if not rows:
        return pd.DataFrame(
            [
                {
                    "Rank": 1,
                    "Stock": "CASH",
                    "Action": "CASH",
                    "Weight %": 1.0,
                    "Allocation TL": REPORT_CAPITAL,
                    "Expected Return %": 0.0,
                    "Expected Return TL": 0.0,
                    "Target Price": None,
                    "Stop Price": None,
                    "Risk/Reward": None,
                    "Conviction": "CASH",
                }
            ]
        )

    summary = pd.DataFrame(rows).sort_values("_attractiveness", ascending=False).head(10).reset_index(drop=True)
    positive = summary["_attractiveness"].clip(lower=0)
    invested_weight = min(len(summary) / max(active_portfolio_size, 1), 1.0)
    if positive.sum() > 0:
        weights = (positive / positive.sum()) * invested_weight
    else:
        weights = pd.Series(invested_weight / len(summary), index=summary.index)

    summary["Rank"] = range(1, len(summary) + 1)
    summary["Weight %"] = weights
    summary["Allocation TL"] = summary["Weight %"] * REPORT_CAPITAL
    summary["Expected Return TL"] = summary["Allocation TL"] * pd.to_numeric(
        summary["Expected Return %"],
        errors="coerce",
    ).fillna(0)

    output = summary[
        [
            "Rank",
            "Stock",
            "Action",
            "Weight %",
            "Allocation TL",
            "Expected Return %",
            "Expected Return TL",
            "Target Price",
            "Stop Price",
            "Risk/Reward",
            "Conviction",
        ]
    ]

    cash_weight = 1.0 - float(output["Weight %"].sum())
    if cash_weight > 0.000001:
        cash_row = pd.DataFrame(
            [
                {
                    "Rank": len(output) + 1,
                    "Stock": "CASH",
                    "Action": "CASH",
                    "Weight %": cash_weight,
                    "Allocation TL": cash_weight * REPORT_CAPITAL,
                    "Expected Return %": 0.0,
                    "Expected Return TL": 0.0,
                    "Target Price": None,
                    "Stop Price": None,
                    "Risk/Reward": None,
                    "Conviction": "CASH",
                }
            ]
        )
        output = pd.concat([output, cash_row], ignore_index=True)

    return output


def _build_portfolio_sheet(current: pd.DataFrame, stock_prices: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    portfolio = current[(current["recommended"] == True) | (current["action"] == "SELL")].copy()
    portfolio = portfolio.sort_values(["recommended", "rank"], ascending=[False, True])

    for _, row in portfolio.iterrows():
        symbol = row["symbol"]
        entry_price = _latest_close(stock_prices, symbol)
        expected_low = row.get("expected_return_low")
        expected_mid = row.get("expected_return_mid")
        expected_high = row.get("expected_return_high")
        quality = _quality_for_row(row, entry_price)
        rows.append(
            {
                "Sıra": int(row["rank"]),
                "Hisse": symbol,
                "Aksiyon": _action_label(row["action"]),
                "Giriş Fiyatı": entry_price,
                "Hedef Alt %": expected_low,
                "Hedef Orta %": expected_mid,
                "Hedef Üst %": expected_high,
                "Hedef Alt Fiyat": _price_target(entry_price, expected_low),
                "Hedef Orta Fiyat": _price_target(entry_price, expected_mid),
                "Hedef Üst Fiyat": _price_target(entry_price, expected_high),
                "Stop / Risk Fiyatı": _risk_price(entry_price, expected_low, row.get("volatility")),
                "Öneri Kalitesi": quality,
                "Güven Puanı": row.get("confidence_score"),
                "Ana Sebep": _main_reason(row),
                "Risk Notu": _risk_note(row),
            }
        )

    return pd.DataFrame(rows)


def _build_top20_sheet(current: pd.DataFrame) -> pd.DataFrame:
    top20 = current.sort_values("rank").head(20).copy()
    return pd.DataFrame(
        {
            "Sıra": top20["rank"].astype(int),
            "Hisse": top20["symbol"],
            "Öneri Kalitesi": top20.apply(_quality_for_row, axis=1),
            "Aksiyon": top20["action"].map(_action_label),
            "1A Momentum": top20["momentum_1m"],
            "3A Momentum": top20["momentum_3m"],
            "6A Momentum": top20["momentum_6m"],
            "Hacim Değişimi": top20["volume_increase"],
            "Trend Durumu": top20["above_ma"].map(_trend_label),
            "Volatilite": top20["volatility"],
            "Beklenen Getiri Orta %": top20["expected_return_mid"],
        }
    )


def _build_action_summary(current: pd.DataFrame) -> pd.DataFrame:
    def symbols_for(action: str) -> str:
        symbols = current.loc[current["action"] == action, "symbol"].tolist()
        return ", ".join(symbols) if symbols else "-"

    excluded = current.loc[current["action"] == "EXCLUDE", "symbol"].tolist()
    min_buy = float(current["min_buy_expected_return"].iloc[0]) if "min_buy_expected_return" in current else 0.10
    opportunity_threshold = (
        float(current["opportunity_threshold"].iloc[0])
        if "opportunity_threshold" in current and pd.notna(current["opportunity_threshold"].iloc[0])
        else min_buy
    )
    qualified_count = int((current["action"] == "BUY").sum())
    return pd.DataFrame(
        [
            {"Kategori": "Alınacaklar", "Hisseler": symbols_for("BUY")},
            {"Kategori": "Tutulacaklar", "Hisseler": symbols_for("HOLD")},
            {"Kategori": "Satılacaklar", "Hisseler": symbols_for("SELL")},
            {"Kategori": "Portföy dışı kalanlar", "Hisseler": ", ".join(excluded) if excluded else "-"},
            {"Kategori": "Nakit kuralı", "Hisseler": f"Beklenen getiri etkin %{opportunity_threshold * 100:.2f} eşiğinin altında ise AL önerisi verilmez; kalan sermaye CASH olur."},
            {"Kategori": "Nitelikli fırsat sayısı", "Hisseler": str(qualified_count)},
        ]
    )


def _load_previous_recommendation(results_path: Path, current: pd.DataFrame) -> pd.DataFrame:
    snapshot_path = results_path.parent / "paper_trading" / "recommendation_snapshots.csv"
    if not snapshot_path.exists():
        return pd.DataFrame()

    snapshots = pd.read_csv(snapshot_path)
    if snapshots.empty or "snapshot_month" not in snapshots:
        return pd.DataFrame()

    current_month = pd.to_datetime(current["snapshot_date"].iloc[0]).to_period("M").strftime("%Y-%m")
    previous = snapshots[snapshots["snapshot_month"] < current_month].copy()
    if previous.empty:
        return pd.DataFrame()

    previous_month = previous["snapshot_month"].max()
    return previous[previous["snapshot_month"] == previous_month].copy()


def _build_portfolio_changes(current: pd.DataFrame, results_path: Path) -> pd.DataFrame:
    previous = _load_previous_recommendation(results_path, current)
    current_actions = current.set_index("symbol")["action"].to_dict()
    previous_actions = previous.set_index("symbol")["action"].to_dict() if not previous.empty else {}
    symbols = sorted(set(current_actions) | set(previous_actions))
    rows = []

    for symbol in symbols:
        current_action = current_actions.get(symbol)
        previous_action = previous_actions.get(symbol)
        if current_action == "BUY" and previous_action is None:
            status = "NEW BUY"
        elif current_action == "BUY" and previous_action == "BUY":
            status = "UNCHANGED"
        elif current_action == "BUY":
            status = "INCREASED"
        elif current_action in {"HOLD", "SELL"} and previous_action == "BUY":
            status = "REDUCED"
        elif current_action is None and previous_action == "BUY":
            status = "REMOVED"
        elif current_action in {"HOLD", "SELL", "EXCLUDE"} and previous_action in {"HOLD", "SELL", "EXCLUDE"}:
            status = "UNCHANGED"
        else:
            status = "REMOVED" if previous_action == "BUY" else "UNCHANGED"
        rows.append({"Hisse": symbol, "Durum": status})

    return pd.DataFrame(rows)


def _build_action_box(executive_summary: pd.DataFrame, current: pd.DataFrame) -> dict[str, str]:
    buys = executive_summary.loc[executive_summary["Action"] == "AL", "Stock"].tolist()
    reduced = current.loc[current["action"] == "HOLD", "symbol"].tolist()
    removed = current.loc[current["action"] == "SELL", "symbol"].tolist()
    cash = _cash_weight(executive_summary)
    return {
        "AL": ", ".join(buys) if buys else "-",
        "AZALT": ", ".join(reduced) if reduced else "-",
        "CIK": ", ".join(removed) if removed else "-",
        "NAKIT": f"%{cash * 100:.0f}",
    }


def _cash_explanation(executive_summary: pd.DataFrame, current: pd.DataFrame) -> list[str]:
    cash = _cash_weight(executive_summary)
    if cash <= 0:
        return []

    active_size = int(current["active_portfolio_size"].iloc[0]) if "active_portfolio_size" in current else 10
    qualified = int((current["action"] == "BUY").sum())
    rejected = max(active_size - qualified, 0)
    return [
        "Neden Nakit Tutuluyor?",
        "- Yeterli sayıda yüksek kaliteli fırsat bulunamadı.",
        "- Opportunity filter aktif.",
        "- Koruma amaçlı nakit oranı artırıldı.",
        f"Qualified opportunities count: {qualified}",
        f"Rejected opportunities count: {rejected}",
        f"Cash allocation %: %{cash * 100:.0f}",
    ]


def _build_market_regime(current: pd.DataFrame) -> pd.DataFrame:
    first = current.iloc[0]
    below_ma200 = _as_bool(first["bist100_below_ma200"])
    risk_status = "Risk OFF" if below_ma200 else "Risk ON"
    explanation = (
        "BIST100 MA200 altında olduğu için savunmacı rejim izleniyor."
        if below_ma200
        else "BIST100 MA200 üzerinde olduğu için risk alma rejimi aktif."
    )

    return pd.DataFrame(
        [
            {"Gösterge": "BIST100 Güncel Kapanış", "Değer": float(first["bist100_close"])},
            {"Gösterge": "BIST100 MA200", "Değer": float(first["bist100_ma200"])},
            {"Gösterge": "Risk ON / Risk OFF", "Değer": risk_status},
            {"Gösterge": "Aktif Model", "Değer": first["active_model"]},
            {"Gösterge": "Güven Puanı", "Değer": float(first["confidence_score"])},
            {"Gösterge": "Açıklama", "Değer": explanation},
        ]
    )


def _build_paper_trade(results_path: Path, stock_prices: dict[str, pd.DataFrame]) -> pd.DataFrame:
    trade_log = _load_csv(results_path / "paper_trade_log.csv")
    history = _load_csv(results_path / "paper_portfolio_history.csv")
    if trade_log.empty:
        active = pd.DataFrame(columns=["symbol", "entry_date", "entry_price", "current_price", "return_pct", "unrealized_pnl"])
    else:
        active = trade_log[trade_log["status"] == "OPEN"].copy()

    latest = history.tail(1).iloc[0] if not history.empty else {}
    portfolio_value = latest.get("portfolio_value", None)
    benchmark_return = latest.get("benchmark_return", None)

    rows = []
    for _, row in active.iterrows():
        symbol = row.get("symbol")
        entry_value = pd.to_numeric(row.get("entry_value"), errors="coerce")
        entry_price = pd.to_numeric(row.get("entry_price"), errors="coerce")
        quantity = pd.to_numeric(row.get("quantity"), errors="coerce")
        current_price = _latest_close(stock_prices, symbol) or entry_price
        return_pct = (current_price / entry_price) - 1 if pd.notna(entry_price) and entry_price else 0.0
        unrealized_pnl = (current_price * quantity) - entry_value if pd.notna(quantity) and pd.notna(entry_value) else 0.0

        rows.append(
            {
                "Aktif Pozisyon": symbol,
                "Giriş Tarihi": row.get("entry_date"),
                "Giriş Fiyatı": entry_price,
                "Güncel Fiyat": current_price,
                "Getiri %": return_pct,
                "Gerçekleşmemiş PnL": unrealized_pnl if pd.notna(unrealized_pnl) else 0.0,
                "Portföy Değeri": portfolio_value,
                "Benchmark Getirisi": benchmark_return,
            }
        )

    return pd.DataFrame(rows)


def _write_markdown(
    path: Path,
    executive_summary: pd.DataFrame,
    portfolio: pd.DataFrame,
    top20: pd.DataFrame,
    action_summary: pd.DataFrame,
    market_regime: pd.DataFrame,
    paper_trade: pd.DataFrame,
    portfolio_changes: pd.DataFrame,
    current: pd.DataFrame,
    report_date: str,
    updated_at: str,
) -> None:
    action_box = _build_action_box(executive_summary, current)
    cash_lines = _cash_explanation(executive_summary, current)
    lines = [
        f"# BIST RankingBot Aylık Yatırımcı Raporu - {report_date}",
        "",
        f"- Rapor Tarihi: {report_date}",
        f"- Son Güncelleme: {updated_at}",
        "",
        "## Yönetici Özeti",
        "",
        "## BU AY YAPILACAKLAR",
        "",
        f"- AL: {action_box['AL']}",
        f"- AZALT: {action_box['AZALT']}",
        f"- ÇIK: {action_box['CIK']}",
        f"- NAKIT: {action_box['NAKIT']}",
        "",
        *cash_lines,
        "",
        executive_summary.to_markdown(index=False, floatfmt=".4f") if not executive_summary.empty else "Veri yok.",
        "",
        "## Portföy Önerisi",
        "",
        portfolio.to_markdown(index=False, floatfmt=".4f") if not portfolio.empty else "Veri yok.",
        "",
        "## İlk 20 Sıralama",
        "",
        top20.to_markdown(index=False, floatfmt=".4f") if not top20.empty else "Veri yok.",
        "",
        "## Al/Sat Özet",
        "",
        action_summary.to_markdown(index=False) if not action_summary.empty else "Veri yok.",
        "",
        "## PORTFOY_DEGISIMI",
        "",
        portfolio_changes.to_markdown(index=False) if not portfolio_changes.empty else "Önceki portföy verisi yok.",
        "",
        "## Piyasa Rejimi",
        "",
        market_regime.to_markdown(index=False) if not market_regime.empty else "Veri yok.",
        "",
        "## Paper Trade",
        "",
        paper_trade.to_markdown(index=False, floatfmt=".4f") if not paper_trade.empty else "Aktif pozisyon yok.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def _style_table_header(sheet, row_number: int, fill: PatternFill, font: Font) -> None:
    for cell in sheet[row_number]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_workbook(
    path: Path,
    portfolio_rows: int,
    current: pd.DataFrame,
    executive_summary: pd.DataFrame,
    report_date: str,
    updated_at: str,
) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    first = current.iloc[0]
    below_ma200 = _as_bool(first["bist100_below_ma200"])
    risk_status = "Risk OFF" if below_ma200 else "Risk ON"
    action_box = _build_action_box(executive_summary, current)
    cash_lines = _cash_explanation(executive_summary, current)
    stale_warning = "Bu rapor güncel olmayabilir." if _report_is_stale(report_date) else ""

    executive_sheet = workbook["YONETICI_OZETI"]
    executive_sheet.insert_rows(1, 18)
    executive_sheet["A1"] = f"BIST RankingBot Aylık Yatırımcı Raporu - {report_date}"
    executive_sheet["A1"].font = Font(bold=True, size=14, color="1F4E78")
    executive_sheet["A2"] = f"Rapor Tarihi: {report_date}"
    executive_sheet["A3"] = f"Son Güncelleme: {updated_at}"
    executive_sheet["A4"] = "100.000 TL Bugün Nasıl Dağıtılır?"
    executive_sheet["A5"] = f"Piyasa Rejimi: {risk_status}"
    executive_sheet["A6"] = f"Güven Puanı: {float(first['confidence_score']):.2f}/100"
    executive_sheet["A7"] = f"Aktif Model: {first['active_model']}"
    executive_sheet["A8"] = "Tablo çekiciliğe göre sıralıdır; yüksek conviction daha güçlü adayları gösterir."

    executive_sheet["A1"] = "BIST RankingBot Yatırımcı Raporu"
    executive_sheet["A4"] = f"Piyasa Rejimi: {risk_status}"
    executive_sheet["A5"] = f"Aktif Model: {first['active_model']}"
    executive_sheet["A6"] = f"Güven Puanı: {float(first['confidence_score']):.2f}/100"
    executive_sheet["A7"] = stale_warning
    executive_sheet["A8"] = ""
    executive_sheet["A9"] = "BU AY YAPILACAKLAR"
    executive_sheet["A9"].font = Font(bold=True, size=13, color="1F4E78")
    executive_sheet["A10"] = f"AL: {action_box['AL']}"
    executive_sheet["A11"] = f"AZALT: {action_box['AZALT']}"
    executive_sheet["A12"] = f"ÇIK: {action_box['CIK']}"
    executive_sheet["A13"] = f"NAKIT: {action_box['NAKIT']}"
    if cash_lines:
        executive_sheet["A15"] = cash_lines[0]
        executive_sheet["A15"].font = Font(bold=True, size=12, color="1F4E78")
        for offset, line in enumerate(cash_lines[1:], start=16):
            executive_sheet[f"A{offset}"] = line
    executive_sheet["A18"] = "Tablo çekiciliğe göre sıralıdır; yüksek conviction daha güçlü adayları gösterir."

    portfolio_sheet = workbook["PORTFOY_ONERISI"]
    portfolio_sheet.insert_rows(1, 6)
    portfolio_sheet["A1"] = f"Rapor Tarihi: {report_date}"
    portfolio_sheet["A2"] = f"Son Güncelleme: {updated_at}"
    portfolio_sheet["A3"] = "Aylık Yatırımcı Özeti"
    portfolio_sheet["A3"].font = Font(bold=True, size=14, color="1F4E78")
    portfolio_sheet["A4"] = f"Toplam önerilen/satılacak satır: {portfolio_rows}"
    portfolio_sheet["A5"] = "Renkler: AL yeşil, TUT sarı, SAT kırmızı."
    portfolio_sheet["A6"] = "Bu rapor araştırma amaçlıdır; otomatik emir veya yatırım tavsiyesi değildir."

    header_rows = {
        "YONETICI_OZETI": 19,
        "PORTFOY_ONERISI": 7,
    }
    for sheet in workbook.worksheets:
        if sheet.title not in header_rows:
            sheet.insert_rows(1, 2)
            sheet["A1"] = f"Rapor Tarihi: {report_date}"
            sheet["A2"] = f"Son Güncelleme: {updated_at}"
            header_rows[sheet.title] = 3

    percent_columns = {
        "Weight %",
        "Expected Return %",
        "Hedef Alt %",
        "Hedef Orta %",
        "Hedef Üst %",
        "1A Momentum",
        "3A Momentum",
        "6A Momentum",
        "Hacim Değişimi",
        "Volatilite",
        "Beklenen Getiri Orta %",
        "Getiri %",
        "Benchmark Getirisi",
    }
    price_columns = {
        "Allocation TL",
        "Expected Return TL",
        "Target Price",
        "Stop Price",
        "Giriş Fiyatı",
        "Hedef Alt Fiyat",
        "Hedef Orta Fiyat",
        "Hedef Üst Fiyat",
        "Stop / Risk Fiyatı",
        "Giriş Fiyatı",
        "Güncel Fiyat",
        "Gerçekleşmemiş PnL",
        "Portföy Değeri",
    }

    for sheet in workbook.worksheets:
        header_row = header_rows[sheet.title]
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _style_table_header(sheet, header_row, header_fill, header_font)

        headers = {cell.value: cell.column for cell in sheet[header_row]}
        for name in percent_columns & headers.keys():
            for cell in sheet.iter_cols(min_col=headers[name], max_col=headers[name], min_row=header_row + 1):
                for item in cell:
                    item.number_format = "0.00%"
        for name in price_columns & headers.keys():
            for cell in sheet.iter_cols(min_col=headers[name], max_col=headers[name], min_row=header_row + 1):
                for item in cell:
                    item.number_format = "#,##0.00"

    for sheet in workbook.worksheets:
        header_row = header_rows[sheet.title]
        headers = {cell.value: cell.column for cell in sheet[header_row]}
        start_row = header_row + 1

        action_col = headers.get("Aksiyon")
        if sheet.title == "YONETICI_OZETI":
            action_col = headers.get("Action")
        rank_col = headers.get("Sıra")
        if sheet.title == "YONETICI_OZETI":
            rank_col = headers.get("Rank")
        if action_col:
            for row in range(start_row, sheet.max_row + 1):
                action = sheet.cell(row=row, column=action_col).value
                fill_color = ACTION_FILLS.get(action)
                if fill_color:
                    for cell in sheet[row]:
                        cell.fill = PatternFill("solid", fgColor=fill_color)
        if rank_col:
            for row in range(start_row, min(sheet.max_row, start_row + 9) + 1):
                for cell in sheet[row]:
                    cell.font = Font(bold=True)

    conviction_col = {cell.value: cell.column for cell in executive_sheet[header_rows["YONETICI_OZETI"]]}.get("Conviction")
    conviction_fills = {
        "VERY HIGH": "C6E0B4",
        "HIGH": "D9EAD3",
        "MEDIUM": "FFF2CC",
        "LOW": "FCE4D6",
    }
    if conviction_col:
        for row in range(header_rows["YONETICI_OZETI"] + 1, executive_sheet.max_row + 1):
            conviction = executive_sheet.cell(row=row, column=conviction_col).value
            fill_color = conviction_fills.get(conviction)
            if fill_color:
                executive_sheet.cell(row=row, column=conviction_col).fill = PatternFill("solid", fgColor=fill_color)

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)

    workbook.save(path)


def generate_investor_report(stock_prices: dict[str, pd.DataFrame], results_dir: str) -> tuple[Path, Path]:
    """Generate human-readable monthly investor Excel and Markdown reports."""
    results_path = Path(results_dir)
    current_path = results_path / "current_month_portfolio.csv"
    if not current_path.exists():
        raise FileNotFoundError("current_month_portfolio.csv must exist before investor report generation.")

    current = pd.read_csv(current_path)
    executive_summary = _build_executive_summary(current, stock_prices)
    portfolio = _build_portfolio_sheet(current, stock_prices)
    top20 = _build_top20_sheet(current)
    action_summary = _build_action_summary(current)
    market_regime = _build_market_regime(current)
    paper_trade = _build_paper_trade(results_path, stock_prices)
    portfolio_changes = _build_portfolio_changes(current, results_path)
    report_date = _report_date_from_current(current)
    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    xlsx_path = results_path / "monthly_investor_report.xlsx"
    md_path = results_path / "monthly_investor_report.md"
    dated_xlsx_path = results_path / f"monthly_investor_report_{report_date}.xlsx"
    dated_md_path = results_path / f"monthly_investor_report_{report_date}.md"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        executive_summary.to_excel(writer, sheet_name="YONETICI_OZETI", index=False)
        portfolio.to_excel(writer, sheet_name="PORTFOY_ONERISI", index=False)
        top20.to_excel(writer, sheet_name="ILK_20_SIRALAMA", index=False)
        action_summary.to_excel(writer, sheet_name="AL_SAT_OZET", index=False)
        portfolio_changes.to_excel(writer, sheet_name="PORTFOY_DEGISIMI", index=False)
        market_regime.to_excel(writer, sheet_name="PIYASA_REJIMI", index=False)
        paper_trade.to_excel(writer, sheet_name="PAPER_TRADE", index=False)

    _style_workbook(xlsx_path, len(portfolio), current, executive_summary, report_date, updated_at)
    _write_markdown(
        md_path,
        executive_summary,
        portfolio,
        top20,
        action_summary,
        market_regime,
        paper_trade,
        portfolio_changes,
        current,
        report_date,
        updated_at,
    )
    copyfile(xlsx_path, dated_xlsx_path)
    copyfile(md_path, dated_md_path)
    return xlsx_path, md_path
