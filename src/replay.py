from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.backtester import run_backtests
from src.current_portfolio import generate_current_month_portfolio
from src.ranking import build_monthly_rankings
from src.selection_exports import build_factor_breakdown


def _price_on_or_after(prices: pd.DataFrame, date: pd.Timestamp) -> tuple[pd.Timestamp, float] | tuple[None, None]:
    close = prices["Close"].dropna().sort_index()
    future = close[close.index >= date]
    if future.empty:
        return None, None
    return pd.Timestamp(future.index[0]), float(future.iloc[0])


def _exit_price_near_target(
    prices: pd.DataFrame,
    entry_date: pd.Timestamp,
    target_date: pd.Timestamp,
) -> tuple[pd.Timestamp, float] | tuple[None, None]:
    close = prices["Close"].dropna().sort_index()
    future = close[close.index >= target_date]
    if not future.empty:
        return pd.Timestamp(future.index[0]), float(future.iloc[0])

    fallback = close[(close.index > entry_date) & (close.index <= target_date)]
    if fallback.empty:
        return None, None
    return pd.Timestamp(fallback.index[-1]), float(fallback.iloc[-1])


def _return_between(prices: pd.DataFrame, entry_date: pd.Timestamp, exit_date: pd.Timestamp) -> float | None:
    _, entry = _price_on_or_after(prices, entry_date)
    _, exit_price = _exit_price_near_target(prices, entry_date, exit_date)
    if entry is None or exit_price is None or entry == 0:
        return None
    return (exit_price / entry) - 1


def _first_trading_day_on_or_after(benchmark_prices: pd.DataFrame, requested_date: str) -> pd.Timestamp:
    requested = pd.Timestamp(requested_date)
    dates = benchmark_prices["Close"].dropna().sort_index().index
    candidates = dates[dates >= requested]
    if candidates.empty:
        raise ValueError(f"No benchmark trading day exists on or after replay date {requested_date}.")
    return pd.Timestamp(candidates[0])


def _truncate_prices(
    prices: dict[str, pd.DataFrame],
    trading_date: pd.Timestamp,
) -> dict[str, pd.DataFrame]:
    return {
        symbol: frame.loc[frame.index <= trading_date].copy()
        for symbol, frame in prices.items()
        if not frame.loc[frame.index <= trading_date].empty
    }


def _build_historical_factor_state(
    stock_prices: dict[str, pd.DataFrame],
    benchmark_prices: pd.DataFrame,
    factor_models: dict[str, dict[str, float]],
    portfolio_sizes: list[int],
    transaction_cost: float,
    illiquid_avg_traded_value_threshold: float,
    speculative_daily_volatility_threshold: float,
) -> tuple[dict[str, pd.DataFrame], pd.DataFrame]:
    rankings_by_model = {}
    all_rankings = []
    all_backtests = []
    all_trades = []

    for model_name, weights in factor_models.items():
        rankings = build_monthly_rankings(stock_prices, weights)
        if rankings.empty:
            continue
        rankings["model"] = model_name
        rankings_by_model[model_name] = rankings
        all_rankings.append(rankings)

        backtest_results, trades = run_backtests(
            stock_prices=stock_prices,
            rankings=rankings,
            portfolio_sizes=portfolio_sizes,
            transaction_cost=transaction_cost,
            model_name=model_name,
        )
        all_backtests.append(backtest_results)
        all_trades.append(trades)

    rankings = pd.concat(all_rankings, ignore_index=True) if all_rankings else pd.DataFrame()
    backtest_results = pd.concat(all_backtests, ignore_index=True) if all_backtests else pd.DataFrame()
    trades = pd.concat(all_trades, ignore_index=True) if all_trades else pd.DataFrame()
    factor_breakdown = build_factor_breakdown(
        trades=trades,
        rankings=rankings,
        backtest_results=backtest_results,
        stock_prices=stock_prices,
        benchmark_prices=benchmark_prices,
        illiquid_avg_traded_value_threshold=illiquid_avg_traded_value_threshold,
        speculative_daily_volatility_threshold=speculative_daily_volatility_threshold,
    )
    return rankings_by_model, factor_breakdown


def _selected_positions(
    recommendation: pd.DataFrame,
    full_stock_prices: dict[str, pd.DataFrame],
    trading_date: pd.Timestamp,
) -> pd.DataFrame:
    active_size = int(recommendation["active_portfolio_size"].iloc[0])
    selected = recommendation[recommendation["action"] == "BUY"].copy()
    rows = []
    weight = 1.0 / active_size if active_size else 0.0
    for _, row in selected.iterrows():
        entry_date, entry_price = _price_on_or_after(full_stock_prices[row["symbol"]], trading_date)
        rows.append(
            {
                "Hisse": row["symbol"],
                "Aksiyon": "AL",
                "Portfoy Agirligi": weight,
                "Giris Tarihi": entry_date.date() if entry_date is not None else None,
                "Giris Fiyati": entry_price,
                "Beklenen Getiri Orta %": row.get("expected_return_mid"),
                "Stop / Risk Fiyati": entry_price * (1 + row.get("expected_return_low", 0))
                if entry_price is not None and pd.notna(row.get("expected_return_low"))
                else None,
            }
        )
    return pd.DataFrame(rows)


def _portfolio_return(
    selected: pd.DataFrame,
    full_stock_prices: dict[str, pd.DataFrame],
    trading_date: pd.Timestamp,
    days: int,
) -> float:
    exit_date = trading_date + pd.Timedelta(days=days)
    total = 0.0
    for _, row in selected.iterrows():
        stock_return = _return_between(full_stock_prices[row["Hisse"]], trading_date, exit_date)
        if stock_return is not None:
            total += float(row["Portfoy Agirligi"]) * stock_return
    return total


def _daily_portfolio_curve(
    selected: pd.DataFrame,
    full_stock_prices: dict[str, pd.DataFrame],
    trading_date: pd.Timestamp,
    exit_date: pd.Timestamp,
    cash_weight: float,
) -> pd.Series:
    curves = []
    for _, row in selected.iterrows():
        prices = full_stock_prices[row["Hisse"]]["Close"].dropna().sort_index()
        window = prices[(prices.index >= trading_date) & (prices.index <= exit_date)].copy()
        if window.empty:
            continue
        entry = window.iloc[0]
        if entry == 0:
            continue
        curves.append((window / entry) * float(row["Portfoy Agirligi"]))
    if not curves:
        return pd.Series([1.0], index=[trading_date])
    combined = pd.concat(curves, axis=1).sum(axis=1) + cash_weight
    return combined.sort_index()


def _max_drawdown(curve: pd.Series) -> float:
    if curve.empty:
        return 0.0
    drawdown = (curve / curve.cummax()) - 1
    return float(drawdown.min())


def _stock_results(
    selected: pd.DataFrame,
    full_stock_prices: dict[str, pd.DataFrame],
    trading_date: pd.Timestamp,
    holding_days: int,
) -> pd.DataFrame:
    rows = []
    exit_target = trading_date + pd.Timedelta(days=holding_days)
    for _, row in selected.iterrows():
        exit_date, exit_price = _exit_price_near_target(full_stock_prices[row["Hisse"]], trading_date, exit_target)
        entry_price = row["Giris Fiyati"]
        realized_return = (exit_price / entry_price) - 1 if entry_price and exit_price else None
        rows.append(
            {
                "Hisse": row["Hisse"],
                "Portfoy Agirligi": row["Portfoy Agirligi"],
                "Giris Fiyati": entry_price,
                "Cikis Tarihi": exit_date.date() if exit_date is not None else None,
                "Cikis Fiyati": exit_price,
                "Gerceklesen Getiri %": realized_return,
            }
        )
    return pd.DataFrame(rows)


def _style_replay_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)
    workbook.save(path)


def run_historical_replay(
    stock_prices: dict[str, pd.DataFrame],
    benchmark_prices: pd.DataFrame,
    requested_date: str,
    holding_days: int,
    results_dir: str,
    factor_models: dict[str, dict[str, float]],
    portfolio_sizes: list[int],
    transaction_cost: float,
    min_buy_expected_return: float,
    opportunity_filter_percentile: float,
    illiquid_avg_traded_value_threshold: float,
    speculative_daily_volatility_threshold: float,
) -> dict[str, object]:
    results_path = Path(results_dir)
    results_path.mkdir(exist_ok=True)

    trading_date = _first_trading_day_on_or_after(benchmark_prices, requested_date)
    historical_stocks = _truncate_prices(stock_prices, trading_date)
    historical_benchmark = benchmark_prices.loc[benchmark_prices.index <= trading_date].copy()
    rankings_by_model, factor_breakdown = _build_historical_factor_state(
        stock_prices=historical_stocks,
        benchmark_prices=historical_benchmark,
        factor_models=factor_models,
        portfolio_sizes=portfolio_sizes,
        transaction_cost=transaction_cost,
        illiquid_avg_traded_value_threshold=illiquid_avg_traded_value_threshold,
        speculative_daily_volatility_threshold=speculative_daily_volatility_threshold,
    )
    recommendation, _ = generate_current_month_portfolio(
        stock_prices=historical_stocks,
        benchmark_prices=historical_benchmark,
        factor_models=factor_models,
        rankings_by_model=rankings_by_model,
        results_dir=results_dir,
        base_model="volume_heavy",
        base_portfolio_size=10,
        defensive_model="low_volatility",
        defensive_portfolio_size=5,
        min_buy_expected_return=min_buy_expected_return,
        opportunity_filter_percentile=opportunity_filter_percentile,
        factor_breakdown=factor_breakdown,
        write_outputs=False,
    )

    selected = _selected_positions(recommendation, stock_prices, trading_date)
    active_size = int(recommendation["active_portfolio_size"].iloc[0])
    cash_weight = max(1.0 - (len(selected) / active_size if active_size else 0.0), 0.0)
    holding_return = _portfolio_return(selected, stock_prices, trading_date, holding_days)
    week_1_return = _portfolio_return(selected, stock_prices, trading_date, 7)
    week_2_return = _portfolio_return(selected, stock_prices, trading_date, 14)
    month_1_return = _portfolio_return(selected, stock_prices, trading_date, 30)
    bist100_return = _return_between(benchmark_prices, trading_date, trading_date + pd.Timedelta(days=holding_days)) or 0.0
    excess_return = holding_return - bist100_return
    curve = _daily_portfolio_curve(
        selected,
        stock_prices,
        trading_date,
        trading_date + pd.Timedelta(days=holding_days),
        cash_weight,
    )
    stock_results = _stock_results(selected, stock_prices, trading_date, holding_days)
    best_stock = "-"
    worst_stock = "-"
    if not stock_results.empty and stock_results["Gerceklesen Getiri %"].notna().any():
        best_stock = str(stock_results.loc[stock_results["Gerceklesen Getiri %"].idxmax(), "Hisse"])
        worst_stock = str(stock_results.loc[stock_results["Gerceklesen Getiri %"].idxmin(), "Hisse"])

    replay_date = pd.Timestamp(requested_date).strftime("%Y-%m-%d")
    summary = {
        "replay_date": replay_date,
        "actual_trading_date": trading_date.strftime("%Y-%m-%d"),
        "holding_days": holding_days,
        "portfolio_return": holding_return,
        "bist100_return": bist100_return,
        "excess_return_vs_bist100": excess_return,
        "max_drawdown": _max_drawdown(curve),
        "best_selected_stock": best_stock,
        "worst_selected_stock": worst_stock,
        "cash_weight": cash_weight,
        "selected_stock_count": len(selected),
    }

    performance = pd.DataFrame(
        [
            {"Donem": "1 Hafta", "Portfoy Getirisi": week_1_return},
            {"Donem": "2 Hafta", "Portfoy Getirisi": week_2_return},
            {"Donem": "1 Ay", "Portfoy Getirisi": month_1_return},
            {"Donem": f"{holding_days} Gun", "Portfoy Getirisi": holding_return},
        ]
    )
    benchmark_comparison = pd.DataFrame(
        [
            {
                "Portfoy Getirisi": holding_return,
                "BIST100 Getirisi": bist100_return,
                "BIST100 Ustu Getiri": excess_return,
                "Maksimum Dusus": summary["max_drawdown"],
            }
        ]
    )
    overview = pd.DataFrame(
        [
            {"Alan": "Replay Date", "Deger": replay_date},
            {"Alan": "Kullanilan Islem Gunu", "Deger": trading_date.strftime("%Y-%m-%d")},
            {"Alan": "Holding Period", "Deger": holding_days},
            {"Alan": "Alinan Hisseler", "Deger": ", ".join(selected["Hisse"].tolist()) if not selected.empty else "-"},
            {"Alan": "Nakit Orani", "Deger": cash_weight},
            {"Alan": "Portfoy Getirisi", "Deger": holding_return},
            {"Alan": "BIST100 Getirisi", "Deger": bist100_return},
            {"Alan": "BIST100 Ustu Getiri", "Deger": excess_return},
            {"Alan": "En Iyi Hisse", "Deger": best_stock},
            {"Alan": "En Kotu Hisse", "Deger": worst_stock},
        ]
    )

    report_path = results_path / f"replay_{replay_date}_report.md"
    xlsx_path = results_path / f"replay_{replay_date}_portfolio.xlsx"
    summary_path = results_path / "replay_summary.csv"
    report_lines = [
        f"# Historical Replay Report - {replay_date}",
        "",
        f"- Requested replay date: {replay_date}",
        f"- Actual trading date used: {trading_date.strftime('%Y-%m-%d')}",
        f"- Holding period: {holding_days} days",
        "",
        "## Bot O Gun Ne Alirdi?",
        "",
        ", ".join(selected["Hisse"].tolist()) if not selected.empty else "Hisse alinmazdi.",
        "",
        f"Cash allocation: {cash_weight:.2%}",
        "",
        "## Sonra Ne Oldu?",
        "",
        f"- 1 week return: {week_1_return:.2%}",
        f"- 2 week return: {week_2_return:.2%}",
        f"- 1 month return: {month_1_return:.2%}",
        f"- {holding_days} day portfolio return: {holding_return:.2%}",
        f"- BIST100 return: {bist100_return:.2%}",
        f"- Excess return vs BIST100: {excess_return:.2%}",
        "",
        f"Beat BIST100: {'Yes' if excess_return > 0 else 'No'}",
        f"Best stock: {best_stock}",
        f"Worst stock: {worst_stock}",
        "",
        "## Hisse Bazli Sonuc",
        "",
        stock_results.to_markdown(index=False, floatfmt=".4f") if not stock_results.empty else "Secili hisse yok.",
        "",
    ]
    report_path.write_text("\n".join(report_lines), encoding="utf-8")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="REPLAY_OZET", index=False)
        selected.to_excel(writer, sheet_name="PORTFOY_ONERISI", index=False)
        performance.to_excel(writer, sheet_name="GERCEKLESEN_PERFORMANS", index=False)
        benchmark_comparison.to_excel(writer, sheet_name="BIST100_KARSILASTIRMA", index=False)
        stock_results.to_excel(writer, sheet_name="HISSE_BAZLI_SONUC", index=False)
    _style_replay_workbook(xlsx_path)

    summary_row = pd.DataFrame([summary])
    if summary_path.exists():
        existing = pd.read_csv(summary_path)
        existing = existing[
            ~(
                (existing["replay_date"] == replay_date)
                & (existing["actual_trading_date"] == trading_date.strftime("%Y-%m-%d"))
                & (existing["holding_days"] == holding_days)
            )
        ]
        summary_row = pd.concat([existing, summary_row], ignore_index=True)
    summary_row.to_csv(summary_path, index=False)

    return {
        **summary,
        "report_path": report_path,
        "xlsx_path": xlsx_path,
        "summary_path": summary_path,
    }
