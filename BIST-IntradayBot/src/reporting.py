from __future__ import annotations

from pathlib import Path
from shutil import copyfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from src.paper_trader import PaperTradeResult
from src.signal_engine import SignalResult


def _unique_archive_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for idx in range(1, 100):
        candidate = path.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Could not create unique archive path for {path}")


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_idx, column in enumerate(ws.columns, start=1):
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 42)
    wb.save(path)


def build_summary_frame(
    report_date: pd.Timestamp | None,
    interval: str | None,
    market_status: str,
    result: PaperTradeResult,
    warnings: list[str],
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Date", report_date.date().isoformat() if report_date is not None else "N/A"),
            ("Market status", market_status),
            ("Data interval", interval or "N/A"),
            ("Starting capital", config.STARTING_CAPITAL),
            ("Trades taken", len(result.trades)),
            ("Total daily return", result.portfolio_return),
            ("BIST100 daily return", result.benchmark_return),
            ("Excess return", result.excess_return),
            ("Best trade", result.best_trade),
            ("Worst trade", result.worst_trade),
            ("Warning count", len(warnings)),
        ],
        columns=["Metric", "Value"],
    )


def write_reports(
    report_date: pd.Timestamp | None,
    interval: str | None,
    signal_result: SignalResult,
    trade_result: PaperTradeResult,
    warnings: list[str],
) -> tuple[Path, Path]:
    config.REPORTS_DIR.mkdir(exist_ok=True)
    config.ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

    date_text = report_date.date().isoformat() if report_date is not None else pd.Timestamp.today().date().isoformat()
    latest_md = config.REPORTS_DIR / "latest_intraday_report.md"
    latest_xlsx = config.REPORTS_DIR / "latest_intraday_report.xlsx"
    archive_md = _unique_archive_path(config.ARCHIVE_DIR / f"intraday_report_{date_text}.md")
    archive_xlsx = _unique_archive_path(config.ARCHIVE_DIR / f"intraday_report_{date_text}.xlsx")

    summary = build_summary_frame(report_date, interval, signal_result.market_status, trade_result, warnings)
    missed = signal_result.missed_signals.copy()
    if missed.empty:
        missed = pd.DataFrame([{"symbol": "Yok", "reason": "Seçim dışı kalan BUY sinyali yok."}])
    notes = pd.DataFrame(
        {
            "Notes": [
                "Research only: this bot does not place real broker orders.",
                "Paper trading only; no live trading.",
                "Yahoo Finance intraday data can be delayed, sparse, or unavailable for BIST symbols.",
                "No overnight holding: all paper positions are closed at the final available candle.",
            ]
        }
    )
    warnings_frame = pd.DataFrame({"Warning": warnings or ["No warning."]})

    with pd.ExcelWriter(latest_xlsx, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="OZET", index=False)
        trade_result.trades.to_excel(writer, sheet_name="TRADES", index=False)
        signal_result.all_signals.to_excel(writer, sheet_name="SIGNALS", index=False)
        missed.to_excel(writer, sheet_name="MISSED_SIGNALS", index=False)
        warnings_frame.to_excel(writer, sheet_name="WARNINGS", index=False)
        notes.to_excel(writer, sheet_name="NOTES", index=False)
    _style_workbook(latest_xlsx)
    copyfile(latest_xlsx, archive_xlsx)

    trades_md = "No paper trades were taken."
    if not trade_result.trades.empty:
        trades_md = trade_result.trades.to_markdown(index=False)

    missed_md = "No missed BUY signals."
    if not signal_result.missed_signals.empty:
        missed_md = signal_result.missed_signals[["symbol", "score", "reason"]].to_markdown(index=False)

    warning_md = "\n".join(f"- {warning}" for warning in warnings) if warnings else "- No warning."
    md = f"""# BIST Intraday Paper Trading Report

- Date: {date_text}
- Market status: {signal_result.market_status}
- Data interval: {interval or "N/A"}
- Starting capital: {config.STARTING_CAPITAL:,.2f} TL
- Trades taken: {len(trade_result.trades)}
- Total daily return: {trade_result.portfolio_return:.2%}
- BIST100 daily return: {trade_result.benchmark_return:.2%}
- Excess return: {trade_result.excess_return:.2%}
- Best trade: {trade_result.best_trade}
- Worst trade: {trade_result.worst_trade}

## Trades Taken

{trades_md}

## Missed Signals

{missed_md}

## Notes

- Research only. No real broker orders are sent.
- Paper trading only. No live trading.
- Yahoo Finance intraday coverage for BIST can be limited.
- No overnight holding is used in this MVP.

## Warnings

{warning_md}
"""
    latest_md.write_text(md, encoding="utf-8")
    copyfile(latest_md, archive_md)
    return latest_md, latest_xlsx
